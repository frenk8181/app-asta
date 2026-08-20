"""
Logica di gestione dei calciatori: import da xlsx, fasce di prezzo per ruolo,
generazione/rilettura markdown, costruzione della coda di chiamata.
Modulo separato e senza dipendenze da Flask per poter essere testato in isolamento.
"""
import re
import random
import time
import openpyxl

ROLE_PLURAL = {"P": "Portieri", "D": "Difensori", "C": "Centrocampisti", "A": "Attaccanti"}
ROLE_PLURAL_REV = {v: k for k, v in ROLE_PLURAL.items()}
VALID_ROLES = ("P", "D", "C", "A")
FASCE_LABELS = ["Top", "Semi-top", "Buoni", "Low-cost"]
FASCE_PERCENTILI = {"Top": 0.85, "Semi-top": 0.60, "Buoni": 0.30, "Low-cost": 0.0}

DEFAULT_PRICE_BANDS = {r: {"type": "auto"} for r in VALID_ROLES}

ROLE_ALIASES = {
    "p": "P", "por": "P", "portiere": "P",
    "d": "D", "dif": "D", "difensore": "D",
    "c": "C", "cen": "C", "centrocampista": "C",
    "a": "A", "att": "A", "attaccante": "A",
}


def parse_players_xlsx(file_stream):
    """Legge un xlsx e isola SOLO le colonne Nome, Sq., R., QUOT. (più 'Fuori lista' per escludere)."""
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active

    header_row_idx = None
    headers = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        norm = [str(c).strip().lower() if c is not None else "" for c in row]
        if "nome" in norm:
            header_row_idx = row_idx
            headers = {name: idx for idx, name in enumerate(norm) if name}
            break
    if header_row_idx is None:
        raise ValueError("Intestazioni non trovate: serve una colonna 'Nome' nelle prime righe del file")

    def find_col(aliases):
        for a in aliases:
            if a in headers:
                return headers[a]
        return None

    col_nome = find_col(["nome"])
    col_squadra = find_col(["sq.", "squadra", "sq"])
    col_ruolo = find_col(["r.", "ruolo", "r"])
    col_quot = find_col(["quot.", "quotazione", "quot"])
    col_fuorilista = find_col(["fuori lista"])

    mancanti = [n for n, c in [("Nome", col_nome), ("Sq.", col_squadra), ("R.", col_ruolo), ("QUOT.", col_quot)] if c is None]
    if mancanti:
        raise ValueError("Colonne mancanti nel file: " + ", ".join(mancanti))

    players = {}
    next_id = 1
    stats = {"importati": 0, "esclusi_fuori_lista": 0, "esclusi_non_validi": 0}

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row is None or len(row) <= max(col_nome, col_squadra, col_ruolo, col_quot):
            continue
        if col_fuorilista is not None and col_fuorilista < len(row) and row[col_fuorilista]:
            stats["esclusi_fuori_lista"] += 1
            continue

        nome = row[col_nome]
        squadra = row[col_squadra]
        ruolo_raw = row[col_ruolo]
        quot = row[col_quot]

        if not nome or not squadra or not ruolo_raw or quot is None:
            stats["esclusi_non_validi"] += 1
            continue

        ruolo = ROLE_ALIASES.get(str(ruolo_raw).strip().lower())
        if not ruolo:
            stats["esclusi_non_validi"] += 1
            continue

        try:
            quot_val = int(quot)
        except (TypeError, ValueError):
            stats["esclusi_non_validi"] += 1
            continue

        players[str(next_id)] = {
            "nome": str(nome).strip(),
            "squadra_reale": str(squadra).strip(),
            "ruolo": ruolo,
            "quotazione": quot_val,
            "chiamato": False,
            "assegnato_a": None,
            "prezzo": None,
        }
        next_id += 1
        stats["importati"] += 1

    return players, next_id, stats


def get_band_thresholds(ruolo, players_of_role, price_bands_config):
    """Ritorna [(label, soglia_minima), ...] in ordine decrescente, sempre le stesse 4 etichette."""
    cfg = (price_bands_config or {}).get(ruolo, {"type": "auto"})
    if cfg.get("type") == "manual" and isinstance(cfg.get("thresholds"), dict):
        th = cfg["thresholds"]
        return [(label, int(th.get(label, 0))) for label in FASCE_LABELS]

    quots = sorted([p["quotazione"] for p in players_of_role], reverse=True)

    def pct(p):
        if not quots:
            return 0
        idx = min(int(len(quots) * (1 - p)), len(quots) - 1)
        return quots[idx]

    return [(label, pct(FASCE_PERCENTILI[label])) for label in FASCE_LABELS]


def band_label(quot, bands):
    for label, min_q in bands:
        if quot >= min_q:
            return label
    return bands[-1][0] if bands else "N/D"


def build_players_markdown(players_dict, price_bands_config):
    lines = ["# Lista Calciatori", ""]
    for ruolo in VALID_ROLES:
        subset = [p for p in players_dict.values() if p["ruolo"] == ruolo]
        if not subset:
            continue
        lines.append(f"## {ROLE_PLURAL[ruolo]} ({len(subset)})")
        lines.append("")
        bands = get_band_thresholds(ruolo, subset, price_bands_config)
        for label, _min_q in bands:
            group = [p for p in subset if band_label(p["quotazione"], bands) == label]
            group.sort(key=lambda p: (-p["quotazione"], p["nome"]))
            if not group:
                continue
            lines.append(f"### {label}")
            for p in group:
                stato = ""
                if p.get("chiamato"):
                    if p.get("assegnato_a"):
                        stato = f" _(assegnato a {p['assegnato_a']} per {p['prezzo']})_"
                    else:
                        stato = " _(chiamato)_"
                lines.append(f"- {p['nome']} | {p['squadra_reale']} | {p['quotazione']}{stato}")
            lines.append("")
    return "\n".join(lines)


def parse_players_markdown(text, existing=None):
    """Rilegge il markdown e ricostruisce il dizionario giocatori.
    Se un giocatore (nome+squadra) esisteva già, ne mantiene lo stato (chiamato/assegnato/prezzo)."""
    existing = existing or {}
    lookup = {}
    for p in existing.values():
        key = (p["nome"].strip().lower(), p["squadra_reale"].strip().lower())
        lookup[key] = p

    players = {}
    next_id = 1
    current_ruolo = None

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if line.startswith("## "):
            name = line[3:].split("(")[0].strip()
            current_ruolo = ROLE_PLURAL_REV.get(name)
        elif line.startswith("- ") and current_ruolo:
            parts = [x.strip() for x in line[2:].split("|")]
            if len(parts) < 3:
                continue
            nome, squadra = parts[0], parts[1]
            m = re.match(r"^\s*(\d+)", parts[2])
            if not nome or not squadra or not m:
                continue
            quot_val = int(m.group(1))
            key = (nome.lower(), squadra.lower())
            old = lookup.get(key)
            players[str(next_id)] = {
                "nome": nome,
                "squadra_reale": squadra,
                "ruolo": current_ruolo,
                "quotazione": quot_val,
                "chiamato": old["chiamato"] if old else False,
                "assegnato_a": old["assegnato_a"] if old else None,
                "prezzo": old["prezzo"] if old else None,
            }
            next_id += 1

    return players, next_id


def build_call_queue(players_dict, ruolo, fascia, ordine, price_bands_config):
    subset = [(pid, p) for pid, p in players_dict.items() if not p["chiamato"]]

    if ruolo:
        subset = [(pid, p) for pid, p in subset if p["ruolo"] == ruolo]

    if fascia:
        from collections import defaultdict
        by_ruolo = defaultdict(list)
        for pid, p in subset:
            by_ruolo[p["ruolo"]].append(p)
        bands_cache = {}
        for r in by_ruolo:
            all_of_role = [pl for pl in players_dict.values() if pl["ruolo"] == r]
            bands_cache[r] = get_band_thresholds(r, all_of_role, price_bands_config)
        subset = [(pid, p) for pid, p in subset if band_label(p["quotazione"], bands_cache[p["ruolo"]]) == fascia]

    if ordine == "valutazione":
        subset.sort(key=lambda x: -x[1]["quotazione"])
    elif ordine == "nome":
        subset.sort(key=lambda x: x[1]["nome"].lower())
    else:
        random.shuffle(subset)

    return [pid for pid, _p in subset]
